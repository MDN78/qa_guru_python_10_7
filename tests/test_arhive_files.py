import pypdf
from utils.path import path_file
from openpyxl import load_workbook
import csv


def test_read_pdf_file():
    reader = pypdf.PdfReader(path_file.resources_directory + '/' + 'Python Testing with Pytest (Brian Okken).pdf')
    reader.pages[1].extract_text()
    assert "Python Testing with pytest" in reader.pages[1].extract_text()


def test_read_xlsx_file():
    workbook = load_workbook(path_file.resources_directory + '/' + 'file_example_XLSX_50.xlsx')
    sheet = workbook.active
    assert sheet.cell(1, 2).value == 'First Name'
    assert sheet.cell(2, 2).value == 'Dulce'
    assert sheet.cell(2, 3).value == 'Abril'


def test_read_csv_file():
    with open(path_file.resources_directory + '/' + 'catalog_auto_google.csv', 'r') as read:
        reader = csv.DictReader(read, delimiter=",")
        assert [row['id'] for row in reader] == ['K456653443']
